import base64
import io
import struct
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from app.services.wsjf_planner_excel_provisioning import gerar_definicao
from wsjf_workbook_package import (
    COLUNAS,
    PLANILHA,
    TABELA,
    erro_graph_indica_workbook_incompativel,
    gerar_wsjf_xlsx,
    reparar_workbook_wsjf,
    validar_workbook_wsjf,
)

TEMPLATE = Path(__file__).resolve().parents[2] / 'templates' / 'wsjf' / 'WSJF.xlsx.base64'


def _template_bytes() -> bytes:
    return base64.b64decode(TEMPLATE.read_text(encoding='ascii'))


def _corromper_indice(xlsx: bytes) -> bytes:
    """Reproduz o defeito real do template antigo: EOCD apontando fora do lugar.

    Nada no corpo do arquivo muda; so o indice deixa de bater. O Excel de
    computador reconstroi e abre; o Graph recusa com FileCorruptTryRepair.
    """
    dados = bytearray(xlsx)
    fim = dados.rfind(b'PK\x05\x06')
    (offset,) = struct.unpack_from('<I', dados, fim + 16)
    struct.pack_into('<I', dados, fim + 16, offset + 1836)
    return bytes(dados)


def _com_linha(xlsx: bytes) -> tuple[bytes, list]:
    workbook = load_workbook(io.BytesIO(xlsx))
    planilha = workbook.worksheets[0]
    # Sem strings vazias: o Excel nao guarda celula vazia, entao ela voltaria
    # como None e o teste compararia uma diferenca de formato, nao de conteudo.
    linha = ['task-1', 'Demanda A', 'bucket-1', 0.5, 'medium', '{}', '2026-09-01', '2026-09-30']
    linha += ['2026-09-02', 'https://tasks.office.com/x', '2026-09-04']
    linha += ['Não', 'aguardando fornecedor', 'Falar com o time', 'Alto', 'observação humana']
    planilha.append(linha)
    planilha.tables[TABELA].ref = f'A1:{planilha.cell(1, len(COLUNAS)).column_letter}2'
    saida = io.BytesIO()
    workbook.save(saida)
    return saida.getvalue(), linha


def test_template_versionado_e_exatamente_o_gerado_pelo_reqsys():
    gerado = gerar_wsjf_xlsx()
    versionado = _template_bytes()

    with zipfile.ZipFile(io.BytesIO(gerado)) as a, zipfile.ZipFile(io.BytesIO(versionado)) as b:
        assert a.namelist() == b.namelist()
        assert all(a.read(nome) == b.read(nome) for nome in a.namelist())


def test_template_versionado_e_aceito_pelo_motor_excel_do_graph():
    resultado = validar_workbook_wsjf(_template_bytes())

    assert resultado['ok'] is True, resultado['erros']
    assert resultado['tabelas'] == [TABELA]
    assert resultado['colunas'] == list(COLUNAS)


def test_template_versionado_abre_com_openpyxl_como_a_prova_de_negocio_faz():
    workbook = load_workbook(io.BytesIO(_template_bytes()))
    planilha = workbook.worksheets[0]

    assert workbook.sheetnames == [PLANILHA]
    assert TABELA in planilha.tables
    assert [celula.value for celula in planilha[1]] == list(COLUNAS)


def test_colunas_cobrem_todos_os_campos_que_o_fluxo_escreve():
    definicao = gerar_definicao(
        {
            'group_id': 'g',
            'plan_id': 'p',
            'excel_source': 's',
            'excel_drive': 'd',
            'excel_file': 'f',
        }
    )
    escritos: set[str] = set()
    pilha = [definicao['actions']]
    while pilha:
        acoes = pilha.pop()
        for acao in acoes.values():
            item = acao.get('inputs', {}).get('parameters', {}).get('item')
            if isinstance(item, dict):
                escritos.update(item)
            for aninhado in (acao.get('actions'), acao.get('else', {}).get('actions')):
                if isinstance(aninhado, dict):
                    pilha.append(aninhado)

    assert escritos
    assert escritos.issubset(set(COLUNAS))


def test_validacao_recusa_indice_zip_inconsistente():
    resultado = validar_workbook_wsjf(_corromper_indice(gerar_wsjf_xlsx()))

    assert resultado['ok'] is False
    assert resultado['erros'] == ['indice_zip_inconsistente']


def test_validacao_recusa_pacote_sem_docprops():
    origem = gerar_wsjf_xlsx()
    saida = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(origem)) as entrada, zipfile.ZipFile(saida, 'w') as destino:
        for nome in entrada.namelist():
            if nome.startswith('docProps/'):
                continue
            destino.writestr(nome, entrada.read(nome))

    resultado = validar_workbook_wsjf(saida.getvalue())

    assert resultado['ok'] is False
    assert 'parte_ausente:docProps/core.xml' in resultado['erros']
    assert 'parte_ausente:docProps/app.xml' in resultado['erros']


def test_reparo_preserva_linhas_e_campos_locais_quando_o_arquivo_e_legivel():
    origem, linha = _com_linha(gerar_wsjf_xlsx())

    reparo = reparar_workbook_wsjf(origem)

    assert reparo['estrategia'] == 'reescrita_preservando_dados'
    assert reparo['linhas_preservadas'] == 1
    assert reparo['avisos'] == []
    assert validar_workbook_wsjf(reparo['conteudo'])['ok'] is True
    planilha = load_workbook(io.BytesIO(reparo['conteudo'])).worksheets[0]
    preservada = [celula.value for celula in planilha[2]]
    assert preservada == linha
    assert dict(zip(COLUNAS, preservada))['Risco'] == 'Alto'


def test_reparo_acrescenta_a_tabela_sem_descartar_um_arquivo_legivel():
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.title = 'Outra'
    workbook.active['A1'] = 'conteudo que nao pode ser perdido'
    origem = io.BytesIO()
    workbook.save(origem)

    reparo = reparar_workbook_wsjf(origem.getvalue())

    assert reparo['estrategia'] == 'tabela_adicionada'
    assert reparo['avisos'] == []
    assert validar_workbook_wsjf(reparo['conteudo'])['ok'] is True
    resultado = load_workbook(io.BytesIO(reparo['conteudo']))
    assert resultado.sheetnames == ['Outra', PLANILHA]
    assert resultado['Outra']['A1'].value == 'conteudo que nao pode ser perdido'


def test_reparo_cai_para_o_template_e_avisa_quando_o_arquivo_e_ilegivel():
    reparo = reparar_workbook_wsjf(_corromper_indice(gerar_wsjf_xlsx()))

    assert reparo['estrategia'] == 'template_canonico'
    assert reparo['linhas_preservadas'] == 0
    assert reparo['avisos'] == ['arquivo_atual_ilegivel:ValueError']
    assert validar_workbook_wsjf(reparo['conteudo'])['ok'] is True


def test_reparo_sem_arquivo_anterior_entrega_o_template_canonico():
    reparo = reparar_workbook_wsjf(None)

    assert reparo['estrategia'] == 'template_canonico'
    assert reparo['conteudo'] == gerar_wsjf_xlsx()


def test_so_recusa_de_pacote_conta_como_workbook_incompativel():
    recusa = {'error': {'code': 'unsupportedWorkbook', 'message': 'FileCorruptTryRepair'}}

    assert erro_graph_indica_workbook_incompativel(recusa) is True
    assert erro_graph_indica_workbook_incompativel({'error': {'code': 'accessDenied'}}) is False
    assert erro_graph_indica_workbook_incompativel('Bad Gateway') is False
    assert erro_graph_indica_workbook_incompativel(None) is False
