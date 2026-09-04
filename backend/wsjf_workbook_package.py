"""Pacote OOXML canonico do WSJF.xlsx (tabela tbDemandas).

Modulo sem dependencia do FastAPI de proposito: ele e importado tanto pelos
servicos do backend quanto por `scripts/bootstrap_wsjf_m365_dev.py`, que roda
em CI com um conjunto minimo de dependencias. `openpyxl` so e importado sob
demanda, dentro de `reparar_workbook_wsjf`.
"""

from __future__ import annotations

import re
import zipfile
from io import BytesIO
from typing import Any

from copilot_memory_simple_package import gerar_planilha_xlsx

PLANILHA = 'Demandas'
TABELA = 'tbDemandas'

# Ordem canonica de tbDemandas, identica a do WSJF.xlsx original do tenant
# (recuperada de xl/tables/table1.xml do template anterior): campos alimentados
# pelo fluxo Planner -> Excel e, no fim, os campos preenchidos por pessoas, que
# o fluxo nunca sobrescreve.
COLUNAS = (
    'TaskId',
    'Título',
    'Bucket',
    'Progresso',
    'Prioridade',
    'Responsáveis',
    'Início',
    'Vencimento',
    'Última alteração',
    'Link Planner',
    'Sincronizado em',
    'Bloqueado',
    'Descrição do bloqueio',
    'Próxima ação',
    'Risco',
    'Observações',
)

PARTES_OBRIGATORIAS = (
    '[Content_Types].xml',
    '_rels/.rels',
    'xl/workbook.xml',
    'docProps/core.xml',
    'docProps/app.xml',
)

# Codigos que o motor Excel do Microsoft Graph devolve quando recusa o pacote
# inteiro. Sao os unicos que autorizam bloquear a instalacao do fluxo: erros de
# permissao, rede ou 5xx nao dizem nada sobre a validade do arquivo.
CODIGOS_WORKBOOK_INCOMPATIVEL = (
    'unsupportedworkbook',
    'filecorrupttryrepair',
    'invalidworkbook',
    'openworkbookfailed',
)


def gerar_wsjf_xlsx() -> bytes:
    """Gera o WSJF.xlsx canonico e vazio, com tbDemandas e suas colunas."""
    return gerar_planilha_xlsx([(PLANILHA, TABELA, list(COLUNAS))])


def _erros_de_integridade_zip(xlsx: bytes) -> list[str]:
    """Valida o container ZIP com o mesmo rigor do motor Excel do Graph.

    O Excel de computador reconstroi o indice quando ele nao bate com o corpo
    do arquivo ("recuperacao"); o Graph nao — devolve FileCorruptTryRepair e
    recusa o pacote. Por isso aqui a checagem e estrita: indice consistente
    (primeiro cabecalho local no offset 0) e CRC de todas as partes.
    """
    if not zipfile.is_zipfile(BytesIO(xlsx)):
        return ['pacote_nao_e_zip']
    try:
        with zipfile.ZipFile(BytesIO(xlsx)) as archive:
            entradas = archive.infolist()
            if not entradas:
                return ['pacote_zip_vazio']
            if min(item.header_offset for item in entradas) != 0:
                return ['indice_zip_inconsistente']
            corrompida = archive.testzip()
            if corrompida:
                return [f'parte_corrompida:{corrompida}']
    except Exception as exc:  # BadZipFile, zlib.error, OSError de offset invalido
        return [f'pacote_zip_corrompido:{type(exc).__name__}']
    return []


def _tabelas_do_pacote(archive: zipfile.ZipFile) -> dict[str, list[str]]:
    tabelas: dict[str, list[str]] = {}
    for nome in archive.namelist():
        if not nome.startswith('xl/tables/') or not nome.endswith('.xml'):
            continue
        texto = archive.read(nome).decode('utf-8', errors='replace')
        match = re.search(r'displayName="([^"]+)"', texto) or re.search(r'\bname="([^"]+)"', texto)
        if not match:
            continue
        tabelas[match.group(1)] = re.findall(r'<tableColumn[^>]*\bname="([^"]*)"', texto)
    return tabelas


def validar_workbook_wsjf(xlsx: bytes) -> dict[str, Any]:
    """Diz se o pacote seria aceito pelo motor Excel do Graph/Power Automate."""
    erros = _erros_de_integridade_zip(xlsx)
    if erros:
        return {'ok': False, 'erros': erros, 'tabelas': [], 'colunas': []}

    with zipfile.ZipFile(BytesIO(xlsx)) as archive:
        nomes = set(archive.namelist())
        erros += [f'parte_ausente:{parte}' for parte in PARTES_OBRIGATORIAS if parte not in nomes]
        content_types = archive.read('[Content_Types].xml').decode('utf-8', errors='replace') if '[Content_Types].xml' in nomes else ''
        for parte in ('/docProps/core.xml', '/docProps/app.xml'):
            if parte not in content_types:
                erros.append(f'content_type_ausente:{parte}')
        tabelas = _tabelas_do_pacote(archive)

    colunas = tabelas.get(TABELA, [])
    if TABELA not in tabelas:
        erros.append(f'tabela_ausente:{TABELA}')
    else:
        faltando = [coluna for coluna in COLUNAS if coluna not in colunas]
        if faltando:
            erros.append(f'colunas_ausentes:{faltando}')
    return {'ok': not erros, 'erros': erros, 'tabelas': sorted(tabelas), 'colunas': colunas}


def _linhas_da_tabela(tabela: Any) -> int:
    from openpyxl.utils.cell import range_boundaries

    _, primeira, _, ultima = range_boundaries(tabela.ref)
    return max(ultima - primeira, 0)


def _acrescentar_tabela(workbook: Any) -> None:
    """Cria a aba de tbDemandas sem tocar no resto de um arquivo legivel.

    Melhor do que substituir o arquivo inteiro pelo template: o que ja estava
    la — outras abas, outras tabelas — continua intacto.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    nome = PLANILHA if PLANILHA not in workbook.sheetnames else f'{PLANILHA} (ReqSys)'
    planilha = workbook.create_sheet(nome)
    planilha.append(list(COLUNAS))
    tabela = Table(displayName=TABELA, ref=f'A1:{get_column_letter(len(COLUNAS))}1')
    tabela.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    planilha.add_table(tabela)


def reparar_workbook_wsjf(atual: bytes | None = None) -> dict[str, Any]:
    """Devolve um WSJF.xlsx aceito pelo Graph, preservando dados quando possivel.

    Reescrever o arquivo atual com openpyxl mantem linhas, tipos e campos
    locais (Risco, Proxima acao, Observacoes) e ja produz um pacote completo,
    com docProps; se o arquivo for legivel mas nao tiver tbDemandas, a aba e
    acrescentada em vez de o arquivo ser trocado. Quando o arquivo atual nao pode
    ser lido — o caso do WSJF.xlsx gerado pelo template quebrado — nao ha o que
    preservar e a unica saida e o template canonico vazio; isso e reportado em
    `avisos`, nunca silenciado.
    """
    avisos: list[str] = []
    if atual:
        # Fora do try de proposito: sem openpyxl nao ha como preservar os dados,
        # e cair calado no template vazio apagaria o arquivo do tenant. Falta da
        # dependencia e problema de implantacao, nao arquivo corrompido.
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(BytesIO(atual))
            planilha = next((aba for aba in workbook.worksheets if TABELA in aba.tables), None)
            if planilha is None:
                estrategia = 'tabela_adicionada'
                linhas = 0
                _acrescentar_tabela(workbook)
            else:
                estrategia = 'reescrita_preservando_dados'
                linhas = _linhas_da_tabela(planilha.tables[TABELA])
            saida = BytesIO()
            workbook.save(saida)
            conteudo = saida.getvalue()
            validacao = validar_workbook_wsjf(conteudo)
            if validacao['ok']:
                return {
                    'conteudo': conteudo,
                    'estrategia': estrategia,
                    'linhas_preservadas': linhas,
                    'avisos': avisos,
                }
            avisos.append(f"reescrita_invalida:{validacao['erros']}")
        except Exception as exc:
            avisos.append(f'arquivo_atual_ilegivel:{type(exc).__name__}')
    return {
        'conteudo': gerar_wsjf_xlsx(),
        'estrategia': 'template_canonico',
        'linhas_preservadas': 0,
        'avisos': avisos,
    }


def erro_graph_indica_workbook_incompativel(erro: Any) -> bool:
    """True somente para recusa do pacote pelo motor Excel, nao para 403/5xx."""
    if isinstance(erro, dict):
        detalhe = erro.get('error') if isinstance(erro.get('error'), dict) else erro
        texto = ' '.join(
            str(detalhe.get(chave) or '') for chave in ('code', 'message', 'innerError')
        )
    else:
        texto = str(erro or '')
    baixo = texto.lower()
    return any(codigo in baixo for codigo in CODIGOS_WORKBOOK_INCOMPATIVEL)
