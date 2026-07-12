#!/usr/bin/env python3
"""Atualizador de planilhas autocontido.

Suporta CSV, JSON e XLSX (XLSX requer openpyxl). Fornece CLI, interface Tkinter
opcional e geração de um pacote portátil para outro computador.
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import shutil
import tempfile
import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any, Iterable

LOGGER = logging.getLogger("atualizador_planilha")
SUPPORTED_EXTENSIONS = {".csv", ".json", ".xlsx"}


class SpreadsheetError(RuntimeError):
    """Erro funcional com mensagem segura para o usuário."""


@dataclass(frozen=True)
class TransformRule:
    column: str
    operation: str
    value: str = ""

    def apply(self, row: dict[str, Any]) -> dict[str, Any]:
        result = dict(row)
        current = result.get(self.column)
        operation = self.operation.lower().strip()
        if operation == "set":
            result[self.column] = self.value
        elif operation == "trim":
            result[self.column] = "" if current is None else str(current).strip()
        elif operation == "upper":
            result[self.column] = "" if current is None else str(current).upper()
        elif operation == "lower":
            result[self.column] = "" if current is None else str(current).lower()
        elif operation == "replace":
            old, separator, new = self.value.partition("=>")
            if not separator:
                raise SpreadsheetError("A operação replace exige valor no formato antigo=>novo.")
            result[self.column] = "" if current is None else str(current).replace(old, new)
        elif operation == "delete":
            result.pop(self.column, None)
        else:
            raise SpreadsheetError(f"Operação não suportada: {self.operation}")
        return result


def _validate_path(path: Path, must_exist: bool = True) -> None:
    if must_exist and not path.is_file():
        raise SpreadsheetError(f"Arquivo não encontrado: {path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise SpreadsheetError(
            f"Formato não suportado: {path.suffix or '<sem extensão>'}. "
            f"Use: {', '.join(sorted(SUPPORTED_EXTENSIONS))}."
        )


def read_records(path: Path) -> list[dict[str, Any]]:
    _validate_path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if suffix == ".json":
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if isinstance(payload, dict):
            payload = payload.get("records", [payload])
        if not isinstance(payload, list) or any(not isinstance(item, dict) for item in payload):
            raise SpreadsheetError("JSON deve conter um objeto ou uma lista de objetos.")
        return [dict(item) for item in payload]
    return _read_xlsx(path)


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SpreadsheetError("XLSX requer a dependência opcional openpyxl.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    names = [str(value) if value is not None else f"coluna_{index + 1}" for index, value in enumerate(headers)]
    return [dict(zip(names, values)) for values in rows]


def write_records(path: Path, records: list[dict[str, Any]]) -> None:
    _validate_path(path, must_exist=False)
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix == ".json":
        with path.open("w", encoding="utf-8") as handle:
            json.dump(records, handle, ensure_ascii=False, indent=2, default=str)
        return
    fields = _collect_fields(records)
    if suffix == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(records)
        return
    _write_xlsx(path, fields, records)


def _collect_fields(records: Iterable[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    for row in records:
        for key in row:
            if key not in fields:
                fields.append(key)
    return fields


def _write_xlsx(path: Path, fields: list[str], records: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SpreadsheetError("XLSX requer a dependência opcional openpyxl.") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dados"
    sheet.append(fields)
    for record in records:
        sheet.append([record.get(field) for field in fields])
    workbook.save(path)


def transform_file(source: Path, destination: Path, rules: list[TransformRule]) -> int:
    records = read_records(source)
    transformed = records
    for rule in rules:
        transformed = [rule.apply(row) for row in transformed]
    write_records(destination, transformed)
    LOGGER.info("Transformação concluída: %s -> %s", source, destination)
    return len(transformed)


def generate_portable_package(destination: Path) -> Path:
    """Gera ZIP portátil contendo este programa, launcher e documentação."""
    destination = destination.resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    source_file = Path(__file__).resolve()
    with tempfile.TemporaryDirectory(prefix="atualizador-planilha-") as temp_dir:
        root = Path(temp_dir) / "atualizador_planilha_portatil"
        root.mkdir()
        shutil.copy2(source_file, root / "atualizador_planilha.py")
        (root / "executar.bat").write_text(
            "@echo off\npython atualizador_planilha.py --ui\nif errorlevel 1 pause\n",
            encoding="utf-8",
        )
        (root / "executar.sh").write_text(
            "#!/usr/bin/env sh\npython3 atualizador_planilha.py --ui\n",
            encoding="utf-8",
        )
        (root / "requirements-optional.txt").write_text("openpyxl>=3.1,<4\n", encoding="utf-8")
        (root / "README.txt").write_text(
            "Atualizador de Planilha Portátil\n\n"
            "Requisito: Python 3.10 ou superior.\n"
            "CSV e JSON: sem dependências externas.\n"
            "XLSX: execute 'python -m pip install -r requirements-optional.txt'.\n"
            "Windows: executar.bat. Linux/macOS: chmod +x executar.sh && ./executar.sh.\n",
            encoding="utf-8",
        )
        archive = shutil.make_archive(str(destination.with_suffix("")), "zip", root.parent, root.name)
    return Path(archive)


class SpreadsheetApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        root.title("Atualizador de Planilha")
        root.geometry("760x420")
        self.source = tk.StringVar()
        self.destination = tk.StringVar()
        self.column = tk.StringVar()
        self.operation = tk.StringVar(value="trim")
        self.value = tk.StringVar()
        self.status = tk.StringVar(value="Selecione os arquivos e configure a regra.")
        self._build()

    def _build(self) -> None:
        frame = ttk.Frame(self.root, padding=18)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Atualizador de Planilha", font=("Arial", 18, "bold")).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 18))
        self._file_row(frame, 1, "Origem", self.source, self._choose_source)
        self._file_row(frame, 2, "Destino", self.destination, self._choose_destination)
        ttk.Label(frame, text="Coluna").grid(row=3, column=0, sticky="w", pady=8)
        ttk.Entry(frame, textvariable=self.column).grid(row=3, column=1, sticky="ew", pady=8)
        ttk.Label(frame, text="Operação").grid(row=4, column=0, sticky="w", pady=8)
        ttk.Combobox(frame, textvariable=self.operation, values=("trim", "upper", "lower", "set", "replace", "delete"), state="readonly").grid(row=4, column=1, sticky="ew", pady=8)
        ttk.Label(frame, text="Valor").grid(row=5, column=0, sticky="w", pady=8)
        ttk.Entry(frame, textvariable=self.value).grid(row=5, column=1, sticky="ew", pady=8)
        ttk.Button(frame, text="Executar transformação", command=self._execute).grid(row=6, column=1, sticky="ew", pady=(18, 8))
        ttk.Button(frame, text="Gerar pacote portátil", command=self._generate).grid(row=7, column=1, sticky="ew", pady=8)
        ttk.Label(frame, textvariable=self.status, wraplength=680).grid(row=8, column=0, columnspan=3, sticky="w", pady=(20, 0))
        frame.columnconfigure(1, weight=1)

    def _file_row(self, frame: ttk.Frame, row: int, label: str, variable: tk.StringVar, command: Any) -> None:
        ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", pady=8)
        ttk.Entry(frame, textvariable=variable).grid(row=row, column=1, sticky="ew", pady=8)
        ttk.Button(frame, text="Selecionar", command=command).grid(row=row, column=2, padx=(8, 0), pady=8)

    def _choose_source(self) -> None:
        value = filedialog.askopenfilename(filetypes=[("Dados", "*.csv *.json *.xlsx")])
        if value:
            self.source.set(value)

    def _choose_destination(self) -> None:
        value = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv"), ("JSON", "*.json"), ("Excel", "*.xlsx")])
        if value:
            self.destination.set(value)

    def _execute(self) -> None:
        try:
            count = transform_file(Path(self.source.get()), Path(self.destination.get()), [TransformRule(self.column.get(), self.operation.get(), self.value.get())])
            self.status.set(f"Concluído: {count} registro(s) processado(s).")
        except Exception as exc:
            LOGGER.exception("Falha na transformação")
            messagebox.showerror("Erro", str(exc))
            self.status.set("Falha. Consulte a mensagem apresentada.")

    def _generate(self) -> None:
        value = filedialog.asksaveasfilename(defaultextension=".zip", filetypes=[("ZIP", "*.zip")])
        if not value:
            return
        try:
            archive = generate_portable_package(Path(value))
            self.status.set(f"Pacote criado: {archive}")
        except Exception as exc:
            LOGGER.exception("Falha ao gerar pacote")
            messagebox.showerror("Erro", str(exc))


def _parse_rule(raw: str) -> TransformRule:
    parts = raw.split(":", 2)
    if len(parts) < 2:
        raise argparse.ArgumentTypeError("Regra deve usar coluna:operação[:valor].")
    return TransformRule(parts[0], parts[1], parts[2] if len(parts) == 3 else "")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Transforma arquivos CSV, JSON e XLSX.")
    parser.add_argument("--source", type=Path)
    parser.add_argument("--destination", type=Path)
    parser.add_argument("--rule", action="append", type=_parse_rule, default=[])
    parser.add_argument("--generate-package", type=Path)
    parser.add_argument("--ui", action="store_true")
    parser.add_argument("--log-level", default="INFO", choices=("DEBUG", "INFO", "WARNING", "ERROR"))
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    logging.basicConfig(level=getattr(logging, args.log_level), format="%(asctime)s %(levelname)s %(name)s %(message)s")
    try:
        if args.generate_package:
            print(generate_portable_package(args.generate_package))
            return 0
        if args.ui:
            root = tk.Tk()
            SpreadsheetApp(root)
            root.mainloop()
            return 0
        if not args.source or not args.destination or not args.rule:
            raise SpreadsheetError("Informe --source, --destination e ao menos uma --rule, ou use --ui.")
        count = transform_file(args.source, args.destination, args.rule)
        print(f"{count} registro(s) processado(s).")
        return 0
    except (SpreadsheetError, OSError, json.JSONDecodeError) as exc:
        LOGGER.error("%s", exc)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
