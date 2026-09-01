#!/usr/bin/env python3
from __future__ import annotations

import argparse
import io
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

GRAPH = "https://graph.microsoft.com/v1.0"
TIMEOUT = 30.0
TABLE = "tbDemandas"
LOCAL_MARKER_FIELDS = ("Risco", "Próxima ação")


def _env(*names: str) -> str:
    for name in names:
        value = os.getenv(name, "").strip()
        if value:
            return value
    return ""


def _required(*names: str) -> str:
    value = _env(*names)
    if not value:
        raise RuntimeError("Variável obrigatória ausente: " + " ou ".join(names))
    return value


def _token(client: httpx.Client) -> str:
    tenant = _required("POWER_PLATFORM_TENANT_ID", "AZURE_TENANT_ID")
    client_id = _required("POWER_PLATFORM_CLIENT_ID", "AZURE_CLIENT_ID")
    secret = _required("POWER_PLATFORM_CLIENT_SECRET", "AZURE_CLIENT_SECRET")
    response = client.post(
        f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token",
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": secret,
            "scope": "https://graph.microsoft.com/.default",
        },
        timeout=TIMEOUT,
    )
    response.raise_for_status()
    token = response.json().get("access_token")
    if not token:
        raise RuntimeError("Microsoft Graph não retornou access_token")
    return str(token)


def _graph(client: httpx.Client, method: str, path: str, token: str, **kwargs: Any) -> httpx.Response:
    headers = {"Authorization": f"Bearer {token}"}
    headers.update(kwargs.pop("headers", {}))
    response = client.request(method, GRAPH + path, headers=headers, timeout=TIMEOUT, **kwargs)
    response.raise_for_status()
    return response


def _discover_target(client: httpx.Client, token: str) -> dict[str, str]:
    explicit = {
        "group_id": _env("WSJF_DEV_GROUP_ID"),
        "plan_id": _env("WSJF_DEV_PLAN_ID"),
        "bucket_id": _env("WSJF_DEV_BUCKET_ID"),
        "drive_id": _env("WSJF_DEV_DRIVE_ID"),
        "file_id": _env("WSJF_DEV_FILE_ID"),
    }
    if all(explicit.values()):
        return explicit

    groups = _graph(
        client,
        "GET",
        "/groups?$top=100&$select=id,displayName,groupTypes",
        token,
    ).json().get("value", [])
    candidates: list[dict[str, str]] = []
    for group in groups:
        if "Unified" not in (group.get("groupTypes") or []):
            continue
        group_id = str(group.get("id") or "")
        try:
            plans = _graph(client, "GET", f"/groups/{group_id}/planner/plans", token).json().get("value", [])
        except httpx.HTTPStatusError:
            continue
        for plan in plans:
            title = str(plan.get("title") or "")
            if "wsjf" not in title.casefold():
                continue
            try:
                drive = _graph(client, "GET", f"/groups/{group_id}/drive?$select=id", token).json()
                children = _graph(
                    client,
                    "GET",
                    f"/groups/{group_id}/drive/root/children?$top=200&$select=id,name,parentReference",
                    token,
                ).json().get("value", [])
            except httpx.HTTPStatusError:
                continue
            matches = [item for item in children if str(item.get("name") or "").casefold() == "wsjf.xlsx"]
            if len(matches) != 1:
                continue
            buckets = _graph(client, "GET", f"/planner/plans/{plan['id']}/buckets", token).json().get("value", [])
            if not buckets:
                continue
            buckets = sorted(buckets, key=lambda item: str(item.get("name") or "").casefold())
            preferred = [b for b in buckets if any(k in str(b.get("name") or "").casefold() for k in ("backlog", "demanda", "entrada"))]
            bucket = preferred[0] if preferred else buckets[0]
            candidates.append(
                {
                    "group_id": group_id,
                    "plan_id": str(plan["id"]),
                    "bucket_id": str(bucket["id"]),
                    "drive_id": str(drive["id"]),
                    "file_id": str(matches[0]["id"]),
                    "group_name": str(group.get("displayName") or ""),
                    "plan_name": title,
                    "bucket_name": str(bucket.get("name") or ""),
                }
            )
    if len(candidates) != 1:
        raise RuntimeError(f"Descoberta WSJF ambígua: esperado 1 alvo real, encontrados {len(candidates)}")
    return candidates[0]


def _download_workbook(client: httpx.Client, token: str, target: dict[str, str]) -> tuple[bytes, str]:
    meta = _graph(client, "GET", f"/drives/{target['drive_id']}/items/{target['file_id']}?$select=id,name,eTag", token).json()
    data = _graph(client, "GET", f"/drives/{target['drive_id']}/items/{target['file_id']}/content", token).content
    return data, str(meta.get("eTag") or "")


def _table_context(data: bytes, task_id: str) -> tuple[Any, Any, Any, list[str], list[int]]:
    wb = load_workbook(io.BytesIO(data))
    for ws in wb.worksheets:
        if TABLE not in ws.tables:
            continue
        table = ws.tables[TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [str(ws.cell(min_row, col).value or "").strip() for col in range(min_col, max_col + 1)]
        if "TaskId" not in headers:
            raise RuntimeError("tbDemandas não contém coluna TaskId")
        task_col = min_col + headers.index("TaskId")
        matches = [row for row in range(min_row + 1, max_row + 1) if str(ws.cell(row, task_col).value or "").strip() == task_id]
        return wb, ws, table, headers, matches
    raise RuntimeError("Tabela tbDemandas não encontrada em WSJF.xlsx")


def _read_task_row(data: bytes, task_id: str) -> dict[str, Any]:
    _, ws, table, headers, matches = _table_context(data, task_id)
    if not matches:
        return {"matching_rows": 0}
    min_col, min_row, _, _ = range_boundaries(table.ref)
    row = matches[0]
    values = {headers[i]: ws.cell(row, min_col + i).value for i in range(len(headers))}
    return {"matching_rows": len(matches), "row": row, "values": values, "header_row": min_row}


def _write_local_markers(data: bytes, task_id: str, marker: str) -> bytes:
    wb, ws, table, headers, matches = _table_context(data, task_id)
    if len(matches) != 1:
        raise RuntimeError(f"Esperada 1 linha antes de gravar marcadores; encontradas {len(matches)}")
    min_col, _, _, _ = range_boundaries(table.ref)
    row = matches[0]
    for field in LOCAL_MARKER_FIELDS:
        if field not in headers:
            raise RuntimeError(f"tbDemandas não contém campo local {field}")
        ws.cell(row, min_col + headers.index(field)).value = marker
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _upload_workbook(client: httpx.Client, token: str, target: dict[str, str], data: bytes, etag: str) -> None:
    headers = {"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    if etag:
        headers["If-Match"] = etag
    _graph(client, "PUT", f"/drives/{target['drive_id']}/items/{target['file_id']}/content", token, headers=headers, content=data)


def _wait_row(client: httpx.Client, token: str, target: dict[str, str], task_id: str, deadline: float, *, title: str | None = None, marker: str | None = None) -> dict[str, Any]:
    last: dict[str, Any] = {"matching_rows": 0}
    while time.time() < deadline:
        try:
            data, _ = _download_workbook(client, token, target)
            last = _read_task_row(data, task_id)
        except Exception:
            # Leitura logo após criação/gravação pode encontrar o arquivo
            # ainda em processamento no SharePoint; trata como não pronto.
            time.sleep(15)
            continue
        values = last.get("values") or {}
        title_ok = title is None or str(values.get("Título") or "") == title
        marker_ok = marker is None or all(str(values.get(field) or "") == marker for field in LOCAL_MARKER_FIELDS)
        if last.get("matching_rows") == 1 and title_ok and marker_ok:
            return last
        time.sleep(15)
    return last


def main() -> int:
    parser = argparse.ArgumentParser(description="Prova real Planner → Power Automate → tbDemandas")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--run-url", required=True)
    parser.add_argument("--wait-seconds", type=int, default=4200)
    args = parser.parse_args()

    marker = f"REQSYS-E2E-{int(time.time())}"
    updated_title = marker + "-ATUALIZADA"
    evidence: dict[str, Any] = {
        "environment": "dev",
        "real": True,
        "mocked": False,
        "simulated": False,
        "source_run_url": args.run_url,
        "captured_at": datetime.now(timezone.utc).isoformat(),
        "planner_task_id": "",
        "excel_task_id": "",
        "excel_matching_rows": 0,
        "local_fields_preserved": False,
        "planner_writeback_detected": True,
        "probe_marker": marker,
    }
    task_id = ""

    try:
        with httpx.Client(follow_redirects=True) as client:
            token = _token(client)
            target = _discover_target(client, token)
            evidence["target"] = {k: target.get(k) for k in ("group_name", "plan_name", "bucket_name") if target.get(k)}

            created = _graph(
                client,
                "POST",
                "/planner/tasks",
                token,
                headers={"Content-Type": "application/json"},
                json={"planId": target["plan_id"], "bucketId": target["bucket_id"], "title": marker},
            ).json()
            task_id = str(created["id"])
            evidence["planner_task_id"] = task_id

            first = _wait_row(client, token, target, task_id, time.time() + args.wait_seconds)
            if first.get("matching_rows") != 1:
                raise RuntimeError("TaskId não apareceu exatamente uma vez em tbDemandas no primeiro ciclo")

            data, etag = _download_workbook(client, token, target)
            _upload_workbook(client, token, target, _write_local_markers(data, task_id, marker), etag)

            task = _graph(client, "GET", f"/planner/tasks/{task_id}", token).json()
            task_etag = str(task.get("@odata.etag") or "")
            if not task_etag:
                raise RuntimeError("Planner não retornou ETag da tarefa")
            _graph(
                client,
                "PATCH",
                f"/planner/tasks/{task_id}",
                token,
                headers={"Content-Type": "application/json", "If-Match": task_etag},
                json={"title": updated_title, "percentComplete": 50},
            )
            after_update = _graph(client, "GET", f"/planner/tasks/{task_id}", token).json()
            etag_after_update = str(after_update.get("@odata.etag") or "")

            second = _wait_row(
                client,
                token,
                target,
                task_id,
                time.time() + args.wait_seconds,
                title=updated_title,
                marker=marker,
            )
            values = second.get("values") or {}
            evidence["excel_task_id"] = str(values.get("TaskId") or "")
            evidence["excel_matching_rows"] = int(second.get("matching_rows") or 0)
            evidence["local_fields_preserved"] = all(str(values.get(field) or "") == marker for field in LOCAL_MARKER_FIELDS)

            planner_after_flow = _graph(client, "GET", f"/planner/tasks/{task_id}", token).json()
            evidence["planner_writeback_detected"] = str(planner_after_flow.get("@odata.etag") or "") != etag_after_update
            evidence["captured_at"] = datetime.now(timezone.utc).isoformat()
    except Exception as exc:
        evidence["probe_error"] = f"{type(exc).__name__}: {str(exc)[:400]}"
    finally:
        if task_id:
            try:
                with httpx.Client(follow_redirects=True) as cleanup:
                    token = _token(cleanup)
                    current = _graph(cleanup, "GET", f"/planner/tasks/{task_id}", token).json()
                    etag = str(current.get("@odata.etag") or "")
                    _graph(cleanup, "DELETE", f"/planner/tasks/{task_id}", token, headers={"If-Match": etag})
                    evidence["planner_cleanup"] = True
            except Exception as exc:
                evidence["planner_cleanup"] = False
                evidence["cleanup_warning"] = f"{type(exc).__name__}: {str(exc)[:240]}"
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(json.dumps(evidence, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    ok = (
        evidence.get("planner_task_id")
        and evidence.get("planner_task_id") == evidence.get("excel_task_id")
        and evidence.get("excel_matching_rows") == 1
        and evidence.get("local_fields_preserved") is True
        and evidence.get("planner_writeback_detected") is False
        and "probe_error" not in evidence
    )
    print(json.dumps({"status": "PASS" if ok else "BLOCKED", "task_id": task_id or None}, ensure_ascii=False, separators=(",", ":")))
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
